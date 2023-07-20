import os
import re
import requests
import tkinter as tk
import pandas as pd
from tkinter import filedialog, messagebox
from functools import partial
from lxml import etree
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import Font, NamedStyle
from docx import Document


def get_word_info(word):
    # 构造请求URL
    url = f'https://www.youdao.com/w/eng/{word}'

    try:
        paraphrase = ""
        data = requests.get(url).text
        html = etree.HTML(data)
        british_pronunciation = html.xpath('//*[@id="phrsListTab"]/h2/div/span[1]/span/text()')[0]
        american_pronunciation = html.xpath('//*[@id="phrsListTab"]/h2/div/span[2]/span/text()')[0]
        li_elements = html.xpath('//*[@id="phrsListTab"]/div/ul')
        for li in li_elements:
            paraphrase = ''.join(li.xpath('.//text()'))
        return british_pronunciation, american_pronunciation, paraphrase
    except Exception as e:
        print(e, word)
        return None


def process_text_files(file_paths):
    for file_path in file_paths:
        content = ""
        # 读取文本文件
        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        elif file_path.endswith('.docx'):
            doc = Document(file_path)
            content = ' '.join([p.text for p in doc.paragraphs])

        # 将内容分隔为单词列表
        words = re.split(r"\b[,.:?!()'\"\s\n\t\r]+?\b|[-_]|\s", content)  # 增加删除空格和-_

        # 全部转换为小写
        lowercase_words = [word.lower() for word in words]

        # 去重
        unique_words = list(set(lowercase_words))

        # 排序
        sorted_words = sorted(unique_words)

        # 清洗
        filtered_words = [word for word in sorted_words if len(word) > 2 and "'" not in word and not re.search(r'[\u4e00-\u9fff]', word) and not re.search(r'\d', word)]  # 删除包含乱码、数字和缩写的单词

        # 导出到Excel
        df = pd.DataFrame(filtered_words, columns=['Words'])
        output_file = file_path.replace('.txt', '.xlsx').replace('.docx', '.xlsx')
        df.to_excel(output_file, index=False)

        # 打开Excel文件
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        worksheet.cell(row=1, column=2, value="British Pronunciation")
        worksheet.cell(row=1, column=3, value="American Pronunciation")
        worksheet.cell(row=1, column=4, value="Paraphrase")

        # 创建标题样式并设置为加粗
        bold_style = NamedStyle(name="bold_style")
        bold_style.font = Font(bold=True)
        worksheet.cell(row=1, column=2).style = bold_style
        worksheet.cell(row=1, column=3).style = bold_style
        worksheet.cell(row=1, column=4).style = bold_style

        # 使用线程池处理请求
        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(get_word_info, word) for word in filtered_words]

            # 遍历每个单元格，获取单词并添加发音和释义
            row_index = 2  # 设置初始单元格
            for future, row in zip(futures, worksheet.iter_rows(min_row=2, max_col=4)):
                word = row[0].value
                word_info = future.result()

                if word_info:
                    british_pronunciation, american_pronunciation, paraphrase = word_info
                    worksheet.cell(row=row_index, column=2).value = british_pronunciation
                    worksheet.cell(row=row_index, column=3).value = american_pronunciation
                    worksheet.cell(row=row_index, column=4).value = paraphrase
                else:
                    # 如果单词发音获取不到，则检查单词是否s,ed,ing结尾，如果是，则去除s,d,ing后再试试
                    if word.endswith(('s', 'ed', 'ing')):
                        word_without_suffix = re.sub(r'(s|d|ing)$', '', word)
                        word_info = get_word_info(word_without_suffix)
                        if word_info:
                            british_pronunciation, american_pronunciation, paraphrase = word_info
                            worksheet.cell(row=row_index, column=2).value = british_pronunciation
                            worksheet.cell(row=row_index, column=3).value = american_pronunciation
                            worksheet.cell(row=row_index, column=4).value = paraphrase

                row_index += 1

        # 保存修改后的Excel文件
        workbook.save(output_file)


def browse_files(file_entry):
    file_paths = filedialog.askopenfilenames(filetypes=[('Text Files', '*.txt'), ('Word Files', '*.docx')], initialdir=os.getcwd())
    if file_paths:
        file_entry.delete(0, tk.END)
        file_entry.insert(tk.END, '\n'.join(file_paths))


def execute_function(file_entry):
    paths = file_entry.get()

    if not paths:
        messagebox.showerror('错误', '请最少选择一个txt或docx文件，可选择多个.')
        return

    execute_button.config(state=tk.DISABLED)

    file_paths = paths.split('\n')
    process_text_files(file_paths)

    messagebox.showinfo('成功', '所有文件已执行完成.')

    execute_button.config(state=tk.NORMAL)


# 创建主窗口
window = tk.Tk()
window.title('英文文章切割为单词 V1.02')
window.configure(bg='sky blue')

# 创建文件浏览小部件
file_label = tk.Label(window, text='请选择一个或多个txt或docx文件:', bg='sky blue')
file_label.pack()

file_entry = tk.Entry(window, width=50)
file_entry.pack()

browse_files_button = tk.Button(window, text='浏览文件', command=partial(browse_files, file_entry))
browse_files_button.pack()

# 创建执行按钮
execute_button = tk.Button(window, text='执行', command=partial(execute_function, file_entry))
execute_button.pack()

# 开始运行
window.mainloop()
