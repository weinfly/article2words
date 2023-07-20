import logging
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from collections import Counter

import docx
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import Font

logging.basicConfig(level=logging.INFO)


class WordInfo:

    def __init__(self, word):
        self.word = word

    def get_info(self):

        url = f'https://www.youdao.com/w/eng/{self.word}'
        try:
            r = requests.get(url)
            html = etree.HTML(r.text)
            british = html.xpath('//span[@class="pronounce"]/span/text()')[0]
            american = html.xpath('//span[@class="pronounce"]/span/text()')[1]
            paraphrase = ''.join(html.xpath('//ul[@id="phrsListTab"]/li//text()'))
            return british, american, paraphrase
        except Exception:
            logging.exception(f'获取单词"{self.word}" 失败')
            return None


def get_word_counts(words):
    return Counter(words)


def clean_words(words):
    cleaned = [word.lower() for word in words if 2 < len(word) < 16 and not re.search(r'[^a-zA-Z\']', word)]
    return sorted(set(cleaned))


def parse_text(text):
    words = re.split(r'[^a-zA-Z\']+', text)
    return words


def process_file(file_path, progress):
    logging.info(f'处理文件:{file_path}')

    book = load_workbook(file_path + '.xlsx')
    sheet = book.active

    words = []
    if file_path.endswith('.txt'):
        with open(file_path) as f:
            words = parse_text(f.read())
    elif file_path.endswith('.docx'):
        doc = docx.Document(file_path)
        words = parse_text(' '.join(p.text for p in doc.paragraphs))

    cleaned_words = clean_words(words)
    counts = get_word_counts(cleaned_words)

    sheet['A1'] = 'Word'
    sheet['B1'] = 'British'
    sheet['C1'] = 'American'
    sheet['D1'] = 'Paraphrase'
    sheet['E1'] = 'Count'

    bold = Font(bold=True)
    sheet['A1'].font = bold
    sheet['B1'].font = bold
    sheet['C1'].font = bold
    sheet['D1'].font = bold
    sheet['E1'].font = bold

    row = 2
    threads = []
    for word in cleaned_words:
        t = threading.Thread(target=query_word, args=(word, row, sheet, counts))
        t.start()
        threads.append(t)
        row += 1

    for t in threads:
        t.join()

    progress.set(progress.get() + 1)
    book.save(file_path + '.xlsx')


def query_word(word, row, sheet, counts):
    word_info = WordInfo(word).get_info()
    if word_info:
        sheet.cell(row, 1, word)
        sheet.cell(row, 2, word_info[0])
        sheet.cell(row, 3, word_info[1])
        sheet.cell(row, 4, word_info[2])
        sheet.cell(row, 5, counts[word])


if __name__ == '__main__':

    root = tk.Tk()
    root.title('单词统计')


    def select_files():
        filepaths = filedialog.askopenfilenames(filetypes=[('Text Files', '*.txt'), ('Word Files', '*.docx')])
        process_selected_files(filepaths)


    def process_selected_files(filepaths):
        if not filepaths:
            messagebox.showerror('错误', '请选择文件')
            return

        max_progress = len(filepaths)
        cur_progress = 0
        progress_var = tk.IntVar(value=cur_progress)
        progress_bar = ttk.Progressbar(root, maximum=max_progress, variable=progress_var)
        progress_bar.pack(fill=tk.X, padx=10, pady=10)

        for f in filepaths:
            process_file(f, progress_var)

        messagebox.showinfo('完成', '处理完成!')
        progress_bar.destroy()


    select_btn = tk.Button(root, text='选择文件', command=select_files)
    select_btn.pack()

    root.mainloop()

logging.info('程序结束')
