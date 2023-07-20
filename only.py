import os
import re
import requests
import tkinter as tk
from tkinter import filedialog, messagebox
from functools import partial
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor


def get_word_info(word):
    # 构造请求URL
    url = f'https://www.youdao.com/w/eng/{word}'

    try:
        paraphrase = ""
        data = requests.get(url).text
        html = etree.HTML(data)

        british_pronunciation = html.xpath('//*[@id="phrsListTab"]/h2/div/span[1]/span/text()')[0]
        american_pronunciation = html.xpath('//*[@id="phrsListTab"]/h2/div/span[2]/span/text()')[0]

        li_elements = html.xpath('//*[@id="phrsListTab"]/div/ul')
        for li in li_elements:
            paraphrase += ''.join(li.xpath('.//text()'))

        return british_pronunciation, american_pronunciation, paraphrase
    except Exception:
        return None


def process_files(file_paths):
    with ProcessPoolExecutor() as executor:
        futures = [executor.submit(process_text_file, file_path) for file_path in file_paths]

        for future in futures:
            future.result()


def browse_files(file_entry):
    file_paths = filedialog.askopenfilenames(filetypes=[('Excel Files', '*.xlsx')],
                                             initialdir=os.getcwd())

    if file_paths:
        file_entry.delete(0, tk.END)
        file_entry.insert(tk.END, '\n'.join(file_paths))


def execute_function(file_entry):
    paths = file_entry.get()

    if not paths:
        messagebox.showerror('错误', '请最少选择一个Excel文件,可选择多个.')
        return

    file_paths = paths.split('\n')

    execute_button.config(state=tk.DISABLED)
    process_files(file_paths)

    messagebox.showinfo('成功', '所有文件已执行完成.')
    execute_button.config(state=tk.NORMAL)


def process_text_file(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    worksheet.cell(row=1, column=2, value="British Pronunciation")
    worksheet.cell(row=1, column=3, value="American Pronunciation")
    worksheet.cell(row=1, column=4, value="Paraphrase")

    # 创建标题样式并设置为加粗
    bold_style = NamedStyle(name="bold_style")
    bold_style.font = Font(bold=True)
    worksheet.cell(row=1, column=2).style = bold_style
    worksheet.cell(row=1, column=3).style = bold_style
    worksheet.cell(row=1, column=4).style = bold_style

    with ThreadPoolExecutor() as executor:
        futures = []

        for row in worksheet.iter_rows(min_row=2):
            word = row[0].value
            future = executor.submit(get_word_info, word)
            futures.append(future)

        for future, row in zip(futures, worksheet.iter_rows(min_row=2)):
            word_info = future.result()

            if word_info:
                british_pron, american_pron, paraphrase = word_info
                worksheet.cell(row=row[0].row, column=2).value = british_pron
                worksheet.cell(row=row[0].row, column=3).value = american_pron
                worksheet.cell(row=row[0].row, column=4).value = paraphrase

            else:
                # 尝试移除词缀
                word_without_suffix = re.sub(r'(s|ed|ing)$', '', word)
                word_info = get_word_info(word_without_suffix)

                if word_info:
                    british_pron, american_pron, paraphrase = word_info
                    worksheet.cell(row=row[0].row, column=2).value = british_pron
                    worksheet.cell(row=row[0].row, column=3).value = american_pron
                    worksheet.cell(row=row[0].row, column=4).value = paraphrase

    workbook.save(file_path)


window = tk.Tk()
window.title('Excel单词本v1.0')
window.configure(bg='sky blue')

file_label = tk.Label(window, text='请选择一个或多个Excel文件:', bg='sky blue')
file_label.pack()

file_entry = tk.Entry(window, width=50)
file_entry.pack()

browse_files_button = tk.Button(window, text='浏览文件', command=partial(browse_files, file_entry))
browse_files_button.pack()

execute_button = tk.Button(window, text='执行', command=partial(execute_function, file_entry))
execute_button.pack()

window.mainloop()
