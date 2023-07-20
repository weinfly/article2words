import os
import re
import requests
import tkinter as tk
import pandas as pd
from tkinter import filedialog, messagebox
from lxml import etree
from openpyxl import load_workbook
from docx import Document
from collections import Counter


class EnglishWordProcessor:

    def __init__(self):
        self.file_paths = []
        self.file_entry = None
        self.execute_btn = None

    @staticmethod
    def get_word_info(word):
        url = f'https://www.youdao.com/w/eng/{word}'

        try:
            paraphrase = ""
            data = requests.get(url).text
            html = etree.HTML(data)
            british_pron = html.xpath('//*[@id="phrsListTab"]/h2/div/span[1]/span/text()')[0]
            american_pron = html.xpath('//*[@id="phrsListTab"]/h2/div/span[2]/span/text()')[0]
            li_elements = html.xpath('//*[@id="phrsListTab"]/div/ul')
            for li in li_elements:
                paraphrase += li.xpath('.//text()')
            return british_pron, american_pron, paraphrase
        except Exception as e:
            print(e, word)
            return None

    def browse_files(self):
        self.file_paths = filedialog.askopenfilenames(filetypes=[('Text Files', '*.txt'), ('Word Files', '*.docx')],
                                                      initialdir=os.getcwd())
        if self.file_paths:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(tk.END, '\n'.join(self.file_paths))

    def execute(self):
        paths = self.file_entry.get()

        if not paths:
            messagebox.showerror('错误', '请最少选择一个txt或docx文件,可选择多个.')
            return

        self.execute_btn.config(state=tk.DISABLED)

        file_paths = paths.split('\n')
        self.process_files(file_paths)

        messagebox.showinfo('成功', '所有文件已执行完成.')

        self.execute_btn.config(state=tk.NORMAL)

    def process_files(self, file_paths):
        for file_path in file_paths:
            content = self.get_file_content(file_path)

            words = self.get_words(content)

            unique_words = self.get_unique_words(words)

            filtered_words = self.filter_words(unique_words)

            word_counts = self.count_words(words)

            df = self.export_to_excel(filtered_words, word_counts)

            workbook = self.load_workbook(df)

            self.format_excel(workbook)

    def get_file_content(self, file_path):
        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        elif file_path.endswith('.docx'):
            doc = Document(file_path)
            return ' '.join([p.text for p in doc.paragraphs])

    def get_words(self, content):
        words = re.split(r'[^a-zA-Z\']+', content)
        return [word.lower() for word in words]

    def get_unique_words(self, words):
        return list(set(words))

    def filter_words(self, words):
        return [word for word in words if len(word) > 2 and "'" not in word
                and not re.search(r'[\u4e00-\u9fff]', word) and not re.search(r'\d', word)]

    def count_words(self, words):
        return Counter(words)

    def export_to_excel(self, words, counts):
        df = pd.DataFrame(words, columns=['Words'])
        df['Word Count'] = df['Words'].map(counts)
        return df

    def load_workbook(self, df):
        file_name = f"{os.path.splitext(self.file_paths[0])[0]}.xlsx"
        df.to_excel(file_name, index=False)
        return load_workbook(file_name)

    def format_excel(self, worksheet):

    def run(self):
        window = tk.Tk()
        window.title('英文文章切割为单词 V1.03')
        window.configure(bg='sky blue')

        file_label = tk.Label(window, text='请选择一个或多个txt或docx文件:', bg='sky blue')
        file_label.pack()

        self.file_entry = tk.Entry(window, width=50)
        self.file_entry.pack()

        browse_btn = tk.Button(window, text='浏览文件', command=self.browse_files)
        browse_btn.pack()

        self.execute_btn = tk.Button(window, text='执行', command=self.execute)
        self.execute_btn.pack()

        window.mainloop()


if __name__ == '__main__':
    processor = EnglishWordProcessor()
    processor.run()
